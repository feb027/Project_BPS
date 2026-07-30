from __future__ import annotations

import re
import uuid
from io import BytesIO
from typing import Any

from django.utils import timezone

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.worksheet.datavalidation import DataValidation

from apps.data.models import Fakta
from apps.katalog.models import Publikasi, Bab, Tabel, KolomTabel
from apps.referensi.models import Indikator, Wilayah, Rincian


def load_workbook_from_upload(upload_file) -> Any:
    if hasattr(upload_file, "seek"):
        upload_file.seek(0)
    return load_workbook(upload_file, read_only=True, data_only=True)


# ── Real kecamatan (IDs 1–39) + official kabupaten (ID 40) ──────────
_KECAMATAN_IDS = list(range(1, 40))
_KABUPATEN_IDS = [40]  # Kabupaten Tasikmalaya (skip typo id=86)


class ManualImportTemplateBuilder:
    MASTER_YEAR = 2026
    SHEET_PREFIX = "T_"          # Table data-sheet prefix

    def __init__(self, publication_year: int, bab_id: int | None = None):
        if publication_year <= self.MASTER_YEAR:
            raise ValueError("publication_year harus lebih besar dari master year 2026")
        self.publication_year = publication_year
        self.master_publikasi = Publikasi.objects.get(tahun_terbit=self.MASTER_YEAR)
        self.bab_id = bab_id

    # ── Public build entry ──────────────────────────────────────────

    def build(self) -> Workbook:
        wb = Workbook()
        self._build_metadata(wb.active)

        self._build_wilayah_reference(wb.create_sheet("_WILAYAH_"))
        self._build_rincian_reference(wb.create_sheet("_RINCIAN_"))

        tabel_qs = self._tabel_qs()
        indikator_ids = self._build_indikator_reference(wb.create_sheet("_INDIKATOR_"), tabel_qs)

        for idx, tabel in enumerate(tabel_qs, start=1):
            ws = wb.create_sheet(self._table_sheet_name(tabel, idx))
            self._build_table_data_sheet(ws, tabel, indikator_ids)

        self._hide_reference_sheets(wb)
        self._lock_template(wb)
        return wb

    # ── Queries ─────────────────────────────────────────────────────

    def _tabel_qs(self):
        """All tables for the selected BAB (or all BABs when no bab_id)."""
        qs = Tabel.objects.filter(
            bab__publikasi=self.master_publikasi,
        )
        if self.bab_id is not None:
            qs = qs.filter(bab_id=self.bab_id)
        return list(qs.order_by("bab__nomor", "pk"))

    def _indikator_from_tables(self, tabel_qs) -> dict[int, Indikator]:
        """All distinct indicators referenced by the given tables."""
        tids = [t.pk for t in tabel_qs]
        koloms = list(
            KolomTabel.objects.filter(tabel_id__in=tids)
            .select_related("indikator")
            .order_by("indikator__nama")
        )
        # Deduplicate by indikator id, preserving first encounter
        seen: dict[int, Indikator] = {}
        for k in koloms:
            if k.indikator_id not in seen:
                seen[k.indikator_id] = k.indikator
        return seen

    # ── _WILAYAH_ sheet (39 kecamatan + 1 kabupaten) ────────────────

    @staticmethod
    def _build_wilayah_reference(ws) -> None:
        ws.append(["wilayah_id", "nama", "jenis"])
        kecamatan = list(
            Wilayah.objects.filter(
                jenis=Wilayah.Jenis.KECAMATAN, id__in=_KECAMATAN_IDS
            ).order_by("id")
        )
        kabupaten = list(
            Wilayah.objects.filter(
                jenis=Wilayah.Jenis.KABUPATEN, id__in=_KABUPATEN_IDS
            ).order_by("id")
        )
        for w in kecamatan + kabupaten:
            ws.append([w.id, w.nama, w.jenis])
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 16
        ws.protection.sheet = True

    # ── _RINCIAN_ sheet ──────────────────────────────────────────────

    @staticmethod
    def _build_rincian_reference(ws) -> None:
        ws.append(["rincian_id", "nama"])
        for r in Rincian.objects.all().order_by("id"):
            ws.append([r.id, r.nama])
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 40
        ws.protection.sheet = True

    # ── _INDIKATOR_ sheet ────────────────────────────────────────────

    def _build_indikator_reference(
        self, ws, tabel_qs
    ) -> dict[int, int]:
        """Return {indikator_id → tabel_id} for the given tables."""
        tids = [t.pk for t in tabel_qs]
        ws.append(["indikator_id", "indikator_nama", "satuan", "tipe_nilai", "tabel_id"])

        rows = list(
            KolomTabel.objects.filter(tabel_id__in=tids)
            .values("indikator_id", "indikator__nama", "indikator__satuan",
                     "indikator__tipe_nilai", "tabel_id")
            .order_by("indikator__nama")
        )
        seen = {}
        for r in rows:
            ws.append([
                r["indikator_id"],
                r["indikator__nama"],
                r["indikator__satuan"] or "",
                r["indikator__tipe_nilai"] or "",
                r["tabel_id"],
            ])
            seen.setdefault(r["indikator_id"], r["tabel_id"])

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 40
        ws.column_dimensions["C"].width = 18
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 12
        ws.protection.sheet = True
        return seen

    # ── Per-table data sheet ────────────────────────────────────────

    def _build_table_data_sheet(
        self, ws, tabel: Tabel, indikator_map: dict[int, int]
    ) -> None:
        koloms = list(
            KolomTabel.objects.filter(tabel=tabel).order_by("urutan")
        )

        # ── Header ──────────────────────────────────────────────────
        if tabel.tipe_baris == Tabel.TipeBaris.KATEGORI:
            headers = ["rincian_id", "nama_rincian"]
        else:
            headers = ["wilayah_id", "nama_wilayah"]

        for k in koloms:
            headers.append(k.indikator.nama)
        ws.append(headers)

        # ── Rows ────────────────────────────────────────────────────
        rows_data = self._get_rows_for_table(tabel)

        for row_idx, row_item in enumerate(rows_data, start=2):
            if tabel.tipe_baris == Tabel.TipeBaris.KATEGORI:
                ws.cell(row=row_idx, column=1, value=row_item["id"])
                ws.cell(row=row_idx, column=2, value=row_item["nama"])
            else:
                ws.cell(row=row_idx, column=1, value=row_item["id"])
                ws.cell(row=row_idx, column=2, value=row_item["nama"])

        # ── Data-validation dropdown for row_id column ──────────────
        last_row = 1 + len(rows_data)
        if tabel.tipe_baris == Tabel.TipeBaris.KATEGORI:
            ref_sheet = "_RINCIAN"
            ref_id_col = "A"
        else:
            ref_sheet = "_WILAYAH"
            ref_id_col = "A"
        dv = DataValidation(
            type="list",
            formula1=f"=_{ref_sheet}!${ref_id_col}$2:${ref_id_col}${last_row}",
            allow_blank=False,
            showErrorMessage=True,
            errorTitle="ID tidak valid",
            error=f"Pilih ID dari sheet _{ref_sheet}.",
        )
        ws.add_data_validation(dv)
        dv.add(f"A2:A{last_row}")

        # ── Style ───────────────────────────────────────────────────
        thin_border = Border(
            left=Side(style="thin", color="C0C0C0"),
            right=Side(style="thin", color="C0C0C0"),
            top=Side(style="thin", color="C0C0C0"),
            bottom=Side(style="thin", color="C0C0C0"),
        )
        header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center_align = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        left_align = Alignment(horizontal="left", vertical="center")

        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
            cell.protection = Protection(locked=True)

        # Style data rows: alternating light gray for readability
        alt_fill = PatternFill("solid", fgColor="F2F7FB")  # very light blue
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for col_idx, cell in enumerate(row):
                cell.border = thin_border
                cell.alignment = center_align
                cell.protection = Protection(locked=False)
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
                # First column (ID) stays bold
                if col_idx == 0:
                    cell.font = Font(bold=True, size=11)
                # Second column (nama) left-aligned
                if col_idx == 1:
                    cell.alignment = Alignment(
                        horizontal="left", vertical="center"
                    )
                # Indicator columns: set number format
                if col_idx >= 2:
                    cell.number_format = "#,##0.00"

        ws.freeze_panes = "A2"

        # ── Auto-fit remaining (indicator) columns ──────────────────
        self._auto_fit_columns(ws)

    # ── Column auto-fit ─────────────────────────────────────────────

    @staticmethod
    def _auto_fit_columns(ws) -> None:
        for col_cells in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col_cells:
                if col_letter is None:
                    col_letter = cell.column_letter
                val = cell.value
                if val is not None:
                    # Rough width: len(str) + 4 padding
                    length = len(str(val)) + 4
                    if length > max_len:
                        max_len = length
            if col_letter and max_len > 0:
                # Cap at 50 to avoid absurdly wide columns
                ws.column_dimensions[col_letter].width = min(max_len, 50)

    # ── Hide reference sheets ───────────────────────────────────────

    @staticmethod
    def _hide_reference_sheets(wb: Workbook) -> None:
        hidden_sheets = {"_METADATA_", "_WILAYAH_", "_RINCIAN_", "_INDIKATOR_"}
        for ws in wb.worksheets:
            if ws.title in hidden_sheets:
                # 'hidden' = unhideable via Excel UI via right-click ➝ Unhide
                # 'veryHidden' = can only be unhidden via VBA
                ws.sheet_state = "veryHidden"

    # ── Row resolution based on tipe_baris ──────────────────────────

    def _get_rows_for_table(self, tabel: Tabel) -> list[dict[str, Any]]:
        if tabel.tipe_baris == Tabel.TipeBaris.KATEGORI:
            return self._rincian_items_for_table(tabel)
        elif tabel.tipe_baris == Tabel.TipeBaris.KABUPATEN:
            return list(
                Wilayah.objects.filter(id__in=_KABUPATEN_IDS)
                .order_by("id")
                .values("id", "nama")
            )
        else:
            # kecamatan (default) — 39 kecamatan + 1 kabupaten total
            rows = list(
                Wilayah.objects.filter(id__in=_KECAMATAN_IDS)
                .order_by("id")
                .values("id", "nama")
            )
            rows += list(
                Wilayah.objects.filter(id__in=_KABUPATEN_IDS)
                .order_by("id")
                .values("id", "nama")
            )
            return rows

    @staticmethod
    def _rincian_items_for_table(tabel: Tabel) -> list[dict[str, Any]]:
        """Return rincian items used by this table in the master publikasi."""
        rincian_ids = (
            Fakta.objects.filter(
                tabel=tabel,
                rincian__isnull=False,
            )
            .values_list("rincian_id", flat=True)
            .distinct()
            .order_by("rincian_id")
        )
        return list(
            Rincian.objects.filter(id__in=list(rincian_ids))
            .order_by("nama")
            .values("id", "nama")
        )

    # ── Metadata ────────────────────────────────────────────────────

    def _build_metadata(self, ws) -> None:
        ws.title = "_METADATA_"
        ws.append(["key", "value"])
        rows = [
            ("master_tahun", str(self.MASTER_YEAR)),
            ("template_version", "1.2"),
            ("publication_year", str(self.publication_year)),
            ("generated_at", timezone.now().isoformat()),
            ("canonical_batch", str(uuid.uuid4())),
        ]
        if self.bab_id is not None:
            try:
                bab = Bab.objects.get(pk=self.bab_id)
                rows.append(("bab_nomor", str(bab.nomor)))
            except Bab.DoesNotExist:
                pass
        for row in rows:
            ws.append(row)
        ws["A1"].font = Font(bold=True)
        ws["B1"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 48
        ws.protection.sheet = True

    # ── Sheet naming ────────────────────────────────────────────────

    @staticmethod
    def _table_sheet_name(tabel: Tabel, index: int) -> str:
        """Excel-compatible per-table sheet name (max 31 chars)."""
        bab = tabel.bab
        nomor = tabel.nomor_tabel or ""
        label = tabel.judul or ""
        raw = f"T_{bab.nomor:02d}{index:02d}_{nomor}_{label}"
        # Replace invalid Excel sheet name characters: \ / ? * [ ] :
        safe = re.sub(r"[\\/?*\[\]:]", "_", raw)
        return safe[:31]

    @staticmethod
    def parse_bab_from_sheet(sheet_name: str) -> int | None:
        """Reverse-lookup bab_nomor from a table sheet name like 'T_0101_...'."""
        if not sheet_name.startswith(ManualImportTemplateBuilder.SHEET_PREFIX):
            return None
        try:
            return int(sheet_name[2:4])
        except (ValueError, IndexError):
            return None

    # ── Lock / serialise ────────────────────────────────────────────

    def _lock_template(self, wb: Workbook) -> None:
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.protection.locked is None:
                        cell.protection = Protection(locked=False)

    def to_bytes(self) -> bytes:
        wb = self.build()
        stream = BytesIO()
        wb.save(stream)
        return stream.getvalue()
