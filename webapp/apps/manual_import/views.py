from __future__ import annotations

import csv
import json
from datetime import datetime
from typing import Any

from django.contrib.auth.decorators import login_required
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import redirect
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST

from rest_framework import status
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAdminUser
from rest_framework.authentication import BasicAuthentication

from openpyxl import load_workbook

from .models import ImportUpload, ImportLog
from .serializers import ImportUploadSerializer
from .services import load_workbook_from_upload
from apps.katalog.models import Bab, Tabel, KolomTabel
from apps.referensi.models import Indikator, Wilayah
from apps.data.models import CanonicalIndicator


MASTER_YEAR = 2026


def _error(reason: str, code: str = "invalid"):
    return {"valid": False, "errors": [{"code": code, "detail": reason}]}


def _ok():
    return {"valid": True, "errors": [], "warnings": []}


def _read_required_sheet(workbook, name: str):
    try:
        return workbook[name]
    except Exception:
        return None


def _read_dataset_sheet(workbook):
    preferred = ["DATA", "Dataset", "Sheet1"]
    for name in preferred:
        if name in workbook.sheetnames:
            return workbook[name]
    return None


def _collect_wilayah_master():
    wilayah_qs = list(
        Wilayah.objects.filter(jenis__in=[Wilayah.Jenis.KABUPATEN, Wilayah.Jenis.KECAMATAN])
        .order_by("-created_at")
    )
    wilayah_map: dict[int, dict[str, Any]] = {}
    for w in wilayah_qs:
        wilayah_map[w.id] = {
            "id": w.id,
            "nama": w.nama,
            "jenis": w.jenis,
        }
    return wilayah_map


def _collect_indikator_master():
    ind_qs = list(
        Indikator.objects.select_related("canonical_indicator")
        .filter(kolomtabel__tabel__bab__publikasi__tahun_terbit=MASTER_YEAR)
        .distinct()
        .order_by("canonical_indicator__code", "nama")
    )
    ind_map: dict[int, dict[str, Any]] = {}
    for ind in ind_qs:
        canonical = getattr(ind, "canonical_indicator", None)
        ind_map[ind.id] = {
            "id": ind.id,
            "nama": ind.nama,
            "satuan": ind.satuan or "",
            "tipe_nilai": getattr(ind, "tipe_nilai", "") or "",
            "canonical_code": getattr(canonical, "code", "") or "",
        }
    return ind_map


def _safe_numeric(value: Any):
    if value in (None, "", "-"):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace(",", ".").replace(" ", "")
    try:
        return float(text)
    except Exception:
        return None


def _extract_upload_payload(workbook, publication_year: int, mode: str):
    ws_wilayah = _read_required_sheet(workbook, "_WILAYAH_")
    ws_indikator = _read_required_sheet(workbook, "_INDIKATOR_")
    ws_meta = _read_required_sheet(workbook, "_METADATA_")
    ws_data = _read_dataset_sheet(workbook)

    if not all([ws_wilayah, ws_indikator, ws_meta, ws_data]):
        return _error("Template Excel tidak lengkap. Butuh _METADATA_, _WILAYAH_, _INDIKATOR_, dan DATA.")

    meta = {}
    for row in ws_meta.iter_rows(min_row=2, values_only=True):
        if row and row[0] is not None:
            meta[row[0]] = str(row[1]).strip() if row[1] is not None else ""

    if meta.get("master_tahun") != str(MASTER_YEAR):
        return _error("Template bukan keluaran master 2026.", "bad_template")

    wilayah_map = _collect_wilayah_master()
    indikator_map = _collect_indikator_master()

    header = []
    headers_row = next(ws_data.iter_rows(min_row=1, max_row=1, values_only=True), [])
    for idx, value in enumerate(headers_row, start=1):
        header.append((idx, (value or "").strip()))

    wilayah_header_idx = next(((i + 1, v) for i, v in header if v == "wilayah_id"), None)
    nama_wilayah_header_idx = next(((i + 1, v) for i, v in header if v == "nama_wilayah"), None)
    if not wilayah_header_idx or not nama_wilayah_header_idx:
        return _error("Header DATA harus memiliki 'wilayah_id' dan 'nama_wilayah'.", "invalid_structure")

    indikator_header_indexes = [
        (idx, label) for idx, label in header if label not in ("", "wilayah_id", "nama_wilayah")
    ]
    if not indikator_header_indexes:
        return _error("Tidak ada indikator di header DATA.", "invalid_structure")

    validation_errors = []
    table_rows = []
    wilayah_aggregate_name = "Kabupaten Tasikmalaya"
    wilayah_set = set(wilayah_map.keys())
    kabupaten_id = None
    for wid, info in wilayah_map.items():
        if info["nama"] == wilayah_aggregate_name and info["jenis"] == Wilayah.Jenis.KABUPATEN:
            kabupaten_id = wid
            break
    kabupaten_ids = {wid for wid, info in wilayah_map.items() if info["jenis"] == Wilayah.Jenis.KABUPATEN}
    kecamatan_ids = {wid for wid, info in wilayah_map.items() if info["jenis"] == Wilayah.Jenis.KECAMATAN}

    used_indikator_ids = set()
    unmatched_labels = []
    for label, idx in [(label, idx) for idx, label in indikator_header_indexes]:
        matched_id = next((i for i, info in indikator_map.items() if info["nama"] == label), None)
        if matched_id is None:
            unmatched_labels.append(label)
        else:
            used_indikator_ids.add(matched_id)

    if unmatched_labels:
        detail = "Indikator tak dikenali: " + ", ".join(unmatched_labels[:10])
        if mode != "strict":
            validation_errors.append({"code": "unknown_indicator", "detail": detail, "warna": "warning"})
        else:
            validation_errors.append({"code": "unknown_indicator", "detail": detail})

    required_wilayah_ids = (kabupaten_ids or set()) | (kecamatan_ids or set())
    present_wilayah_ids = set()
    extra_rows = 0
    missing_kabupaten = False
    duplicate_rows = False

    seen_wilayah = set()
    data_rows = []
    for row in ws_data.iter_rows(min_row=2, values_only=True):
        wilayah_value = row[wilayah_header_idx[0] - 1] if wilayah_header_idx and len(row) >= wilayah_header_idx[0] else None
        if wilayah_value in (None, ""):
            continue
        try:
            wilayah_id = int(wilayah_value)
        except Exception:
            validation_errors.append({"code": "invalid_wilayah", "detail": f"wilayah_id tidak valid: {wilayah_value}"})
            continue

        if wilayah_id not in wilayah_set:
            validation_errors.append({"code": "unknown_wilayah", "detail": f"wilayah_id {wilayah_id} tidak ditemukan di master."})
            continue

        if wilayah_id in seen_wilayah:
            duplicate_rows = True
        seen_wilayah.add(wilayah_id)
        present_wilayah_ids.add(wilayah_id)

        row_data = {
            "wilayah_id": wilayah_id,
            "nama_wilayah": wilayah_map[wilayah_id]["nama"],
            "values": {},
        }
        for idx, label in indikator_header_indexes:
            value = row[idx - 1] if idx <= len(row) else None
            row_data["values"][label] = value
        data_rows.append(row_data)

    if kabupaten_ids and kabupaten_id not in present_wilayah_ids:
        missing_kabupaten = True
    if kecamatan_ids and not kecamatan_ids.issubset(present_wilayah_ids):
        validation_errors.append({"code": "missing_wilayah", "detail": "Sebagian kecamatan tidak ada di DATA."})

    if duplicate_rows:
        validation_errors.append({"code": "duplicate_wilayah", "detail": "Ada duplikat wilayah_id di DATA."})

    report = _ok()
    report["warnings"] = [
        x for x in validation_errors if x.get("code") == "unknown_indicator" and mode != "strict"
    ]
    if missing_kabupaten:
        report["warnings"].append({"code": "missing_kabupaten", "detail": "Baris Kabupaten Tasikmalaya tidak ditemukan."})
    if mode == "strict":
        hard_errors = [x for x in validation_errors if x.get("code") != "unknown_indicator"]
    else:
        hard_errors = [x for x in validation_errors if x.get("code") not in ("unknown_indicator",)]
    if hard_errors:
        report["valid"] = False
        report["errors"] = hard_errors

    summary = {
        "publication_year": publication_year,
        "master_source_year": MASTER_YEAR,
        "header_columns": len(header),
        "indikator_columns": len(indikator_header_indexes),
        "wilayah_present": len(present_wilayah_ids),
        "wilayah_required": len(required_wilayah_ids),
        "data_rows": len(data_rows),
        "unmatched_indicator_labels": unmatched_labels,
    }
    return {
        "valid": report["valid"],
        "errors": report["errors"],
        "warnings": report["warnings"],
        "summary": summary,
        "data_rows": data_rows,
        "wilayah_map": wilayah_map,
        "indikator_map": indikator_map,
        "header": header,
        "indikator_header_indexes": indikator_header_indexes,
        "mode": mode,
    }


@api_view(["POST"])
@authentication_classes([BasicAuthentication])
@permission_classes([IsAdminUser])
def generate_template(request):
    try:
        publication_year = int(request.data.get("publication_year", ""))
    except Exception:
        return Response({"error": " publication_year harus angka."}, status=status.HTTP_400_BAD_REQUEST)

    from apps.manual_import.services import ManualImportTemplateBuilder
    builder = ManualImportTemplateBuilder(publication_year=publication_year)
    workbook = builder.build()

    from openpyxl import Workbook
    workbook.save(f"/tmp/manual_import_template_{publication_year}.xlsx")
    with open(f"/tmp/manual_import_template_{publication_year}.xlsx", "rb") as f:
        data = f.read()

    response = HttpResponse(data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=template_bps_master_{MASTER_YEAR}_{publication_year}.xlsx"
    return response


@api_view(["POST"])
@authentication_classes([BasicAuthentication])
@permission_classes([IsAdminUser])
def upload(request):
    upload_file = request.FILES.get("file")
    if not upload_file:
        return Response({"error": "File Excel wajib diupload."}, status=status.HTTP_400_BAD_REQUEST)

    publication_year = request.data.get("publication_year")
    mode = request.data.get("mode", "strict")
    if publication_year is None:
        return Response({"error": "publication_year wajib diisi."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        publication_year = int(publication_year)
    except Exception:
        return Response({"error": "publication_year harus angka."}, status=status.HTTP_400_BAD_REQUEST)

    upload_obj = ImportUpload(
        publication_year=publication_year,
        master_source_year=MASTER_YEAR,
        mode=mode,
        original_filename=upload_file.name,
        file=upload_file,
        status=ImportUpload.Status.UPLOADED,
    )
    upload_obj.save()

    try:
        workbook = load_workbook_from_upload(upload_file)
        payload = _extract_upload_payload(workbook, publication_year, mode)
    except Exception as e:
        upload_obj.status = ImportUpload.Status.REJECTED
        upload_obj.validation_report = {"valid": False, "errors": [{"detail": str(e)}]}
        upload_obj.processed_at = timezone.now()
        upload_obj.save()
        return Response({"upload_id": upload_obj.id, "error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    upload_obj.validation_report = {
        "valid": payload["valid"],
        "errors": payload["errors"],
        "warnings": payload["warnings"],
        "summary": payload["summary"],
    }
    if payload["valid"]:
        upload_obj.status = ImportUpload.Status.VALIDATED
        upload_obj.preview_summary = {
            "data_rows": len(payload["data_rows"]),
            "indikator_count": len(payload["indikator_header_indexes"]),
        }
    else:
        upload_obj.status = ImportUpload.Status.REJECTED
    upload_obj.processed_at = timezone.now()
    upload_obj.save()

    preview = {
        "upload_id": str(upload_obj.id),
        "publication_year": publication_year,
        "master_source_year": MASTER_YEAR,
        "mode": mode,
        "validation": {
            "is_valid": payload["valid"],
            "errors": payload["errors"],
            "warnings": payload["warnings"],
        },
        "summary": payload["summary"],
        "preview_rows": payload["data_rows"][:50],
        "preview_row_count": len(payload["data_rows"]),
    }
    return Response({"upload_id": upload_obj.id, "preview": preview}, status=status.HTTP_200_OK)


@api_view(["GET"])
@authentication_classes([BasicAuthentication])
@permission_classes([IsAdminUser])
def preview(request, pk: str):
    try:
        upload = ImportUpload.objects.get(pk=pk)
    except ImportUpload.DoesNotExist:
        return Response({"error": "Upload tidak ditemukan."}, status=status.HTTP_404_NOT_FOUND)

    return Response({"preview": upload.preview_summary, "validation": upload.validation_report})


@api_view(["POST"])
@authentication_classes([BasicAuthentication])
@permission_classes([IsAdminUser])
def commit(request, pk: str):
    try:
        upload = ImportUpload.objects.get(pk=pk)
    except ImportUpload.DoesNotExist:
        return Response({"error": "Upload tidak ditemukan."}, status=status.HTTP_404_NOT_FOUND)

    if upload.status != ImportUpload.Status.VALIDATED:
        return Response({"error": "Upload belum valid atau sudah diproses."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        workbook = load_workbook_from_upload(upload.file)
        payload = _extract_upload_payload(workbook, upload.publication_year, upload.mode)

        if not payload["valid"]:
            return Response({"error": "Data tidak valid.", "validation": payload}, status=status.HTTP_400_BAD_REQUEST)
    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    wilayah_map = payload["wilayah_map"]
    indikator_map = payload["indikator_map"]
    data_rows = payload["data_rows"]
    header = payload["header"]
    indikator_header_indexes = payload["indikator_header_indexes"]

    master_publikasi = Publikasi.objects.get(tahun_terbit=MASTER_YEAR)
    master_babs = list(Bab.objects.filter(publikasi=master_publikasi).prefetch_related("tabel_set__kolomtabel_set__indikator"))
    if not master_babs:
        return Response({"error": "Master 2026 belum memiliki struktur Bab/Tabel."}, status=status.HTTP_400_BAD_REQUEST)

    # Placeholder mapping: gunakan Bab master pertama untuk commit awal
    target_bab = master_babs[0]
    target_tabel, _ = Tabel.objects.get_or_create(
        bab=target_bab,
        judul=f"Imported Manual {upload.publication_year}",
        defaults={"tipe_baris": Tabel.TipeBaris.KECAMATAN},
    )

    indikator_ids = [ind for _, label in indikator_header_indexes for ind, info in indikator_map.items() if info["nama"] == label]
    indikator_positions = {label: next(ind for ind, info in indikator_map.items() if info["nama"] == label) for label in [label for _, label in indikator_header_indexes] if any(info["nama"] == label for info in indikator_map.values())}
    created_fakta = 0
    skipped_rows = 0

    from django.db import transaction
    with transaction.atomic():
        for row in data_rows:
            wilayah_id = row["wilayah_id"]
            for label, _ in indikator_header_indexes:
                indikator_id = indikator_positions.get(label)
                if indikator_id is None:
                    skipped_rows += 1
                    continue
                nilai_num = _safe_numeric(row["values"].get(label))
                nilai_teks = None if nilai_num is not None else str(row["values"].get(label)).strip()
                Fakta.objects.create(
                    tabel=target_tabel,
                    wilayah=Wilayah.objects.get(id=wilayah_id),
                    kolom=KolomTabel.objects.get_or_create(tabel=target_tabel, indikator_id=indikator_id)[0],
                    tahun=upload.publication_year,
                    nilai_num=nilai_num,
                    nilai_teks=nilai_teks or "-",
                )
                created_fakta += 1

        upload.status = ImportUpload.Status.COMMITTED
        upload.processed_at = timezone.now()
        upload.save()

        ImportLog.objects.create(
            upload=upload,
            user=request.user,
            publication_year=upload.publication_year,
            master_source_year=MASTER_YEAR,
            mode=upload.mode,
            status=ImportLog.Status.COMMITTED,
            raw_filename=upload.original_filename,
            validation_report=upload.validation_report,
            preview_summary=upload.preview_summary,
            tables_affected=[{"tabel_id": target_tabel.id, "judul": target_tabel.judul}],
            faktas_inserted=created_fakta,
            committed_at=timezone.now(),
        )

    return Response({"status": "committed", "faktas_inserted": created_fakta, "skipped_rows": skipped_rows})


@require_GET
def placeholder(request):
    return JsonResponse({"status": "ok", "phase": "phase1_scaffold"})
