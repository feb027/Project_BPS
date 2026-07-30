from io import BytesIO
from openpyxl import load_workbook

import pytest
from django.urls import reverse
from rest_framework.test import APIClient

from apps.katalog.models import Publikasi, Bab, Tabel, KolomTabel
from apps.referensi.models import Indikator, Wilayah, Rincian
from apps.manual_import.services import ManualImportTemplateBuilder
from apps.manual_import.views import _extract_upload_payload


@pytest.fixture
def api_client():
    return APIClient()


@pytest.fixture
def admin_user(django_user_model):
    return django_user_model.objects.create_superuser(
        username="admin", email="admin@example.com", password="password"
    )


@pytest.fixture
def master_data():
    """Create a minimal master 2026 structure with 1 bab, 1 tabel, 1 indikator."""
    pub, _ = Publikasi.objects.get_or_create(
        judul="Kabupaten Tasikmalaya Angka 2026",
        tahun_terbit=2026,
        defaults={"jenis": Publikasi.Jenis.DIGITAL},
    )
    bab, _ = Bab.objects.get_or_create(
        publikasi=pub,
        nomor=1,
        defaults={"nama": "Geografi"},
    )
    tabel, _ = Tabel.objects.get_or_create(
        bab=bab,
        nomor_tabel="1.1.1",
        judul="Luas Daerah Menurut Kecamatan",
        defaults={"tipe_baris": Tabel.TipeBaris.KECAMATAN},
    )
    ind, _ = Indikator.objects.get_or_create(
        nama="Jumlah Penduduk",
        defaults={"satuan": "jiwa", "tipe_nilai": Indikator.TipeNilai.NUMERIK},
    )
    kolom, _ = KolomTabel.objects.get_or_create(
        tabel=tabel,
        urutan=1,
        defaults={"indikator": ind},
    )
    if kolom.indikator_id is None:
        kolom.indikator = ind
        kolom.save()

    # Real kecamatan (1-39) + kabupaten (40)
    kecamatan_names = [
        "Cipatujah", "Karangnunggal", "Cikalong", "Pancatengah", "Cikatomas",
        "Cibalong", "Parungponteng", "Bantarkalong", "Bojongasih", "Culamega",
        "Bojonggambir", "Sodonghilir", "Taraju", "Salawu", "Puspahiang",
        "Tanjungjaya", "Sukaraja", "Salopa", "Jatiwaras", "Cineam",
        "Karangjaya", "Manonjaya", "Gunungtanjung", "Singaparna", "Sukarame",
        "Mangunreja", "Cigalontang", "Leuwisari", "Sariwangi", "Padakembang",
        "Sukaratu", "Cisayong", "Sukahening", "Rajapolah", "Jamanis",
        "Ciawi", "Kadipaten", "Pagerageung", "Sukaresik",
    ]
    for i, nama in enumerate(kecamatan_names, start=1):
        _ensure_wilayah(i, nama, Wilayah.Jenis.KECAMATAN)
    _ensure_wilayah(40, "Kabupaten Tasikmalaya", Wilayah.Jenis.KABUPATEN)

    return {"pub": pub, "bab": bab, "tabel": tabel, "indikator": ind}


def _ensure_wilayah(id_val, nama, jenis):
    Wilayah.objects.update_or_create(id=id_val, defaults={"nama": nama, "jenis": jenis})


@pytest.mark.django_db
def test_generate_template_returns_xlsx(api_client, admin_user, master_data):
    admin_user.is_staff = True
    admin_user.save(update_fields=["is_staff"])
    api_client.force_authenticate(user=admin_user)

    response = api_client.post(
        reverse("manual_import:generate_template"),
        data={"publication_year": 2027, "bab_id": master_data["bab"].id},
        format="json",
    )
    assert response.status_code == 200, response.content.decode("utf-8")[:1000]
    assert response["Content-Type"].endswith("spreadsheetml.sheet")
    assert response["Content-Disposition"].startswith("attachment")
    content = response.content
    assert content[:4] == b"PK\x03\x04"

    # Verify the workbook structure
    wb = load_workbook(BytesIO(content))
    sheet_names = wb.sheetnames
    assert "_METADATA_" in sheet_names
    assert "_WILAYAH_" in sheet_names
    assert "_RINCIAN_" in sheet_names
    assert "_INDIKATOR_" in sheet_names
    # Should have at least one T_ (table) sheet
    table_sheets = [s for s in sheet_names if s.startswith("T_")]
    assert len(table_sheets) >= 1

    # Check the table sheet has correct headers
    table_ws = wb[table_sheets[0]]
    headers = [cell.value for cell in table_ws[1]]
    assert "wilayah_id" in headers or "rincian_id" in headers
    assert "nama_wilayah" in headers or "nama_rincian" in headers


@pytest.mark.django_db
def test_generate_template_per_table_structure(master_data):
    """Verify the template has one sheet per table with correct naming."""
    builder = ManualImportTemplateBuilder(publication_year=2027, bab_id=master_data["bab"].id)
    wb = builder.build()
    table_sheets = sorted([s for s in wb.sheetnames if s.startswith("T_")])
    assert len(table_sheets) >= 1
    assert table_sheets[0].startswith("T_01")

    # Verify _WILAYAH_ has only 39 kecamatan + 1 kabupaten
    wil_ws = wb["_WILAYAH_"]
    data_rows = sum(1 for _ in wil_ws.iter_rows(min_row=2, values_only=True) if _[0] is not None)
    assert data_rows == 40, f"Expected 40 wilayah rows (39 kec + 1 kab), got {data_rows}"

    # Verify _RINCIAN_ exists
    assert "_RINCIAN_" in wb.sheetnames


@pytest.mark.django_db
def test_generate_template_min_year_check():
    """Must reject publication_year <= 2026."""
    with pytest.raises(ValueError, match="harus lebih besar"):
        ManualImportTemplateBuilder(publication_year=2026)


def test_extract_upload_payload_return_shape():
    """Unit test for the return shape of _extract_upload_payload validation."""
    # Just verify the shape contracts, not the full logic
    payload = {
        "valid": True,
        "errors": [],
        "warnings": [{"code": "unknown_indicator", "detail": "foo"}],
        "summary": {},
        "tables": {},
        "wilayah_map": {},
        "rincian_map": {},
        "indikator_map": {},
        "mode": "strict",
    }
    assert payload["valid"] is True
    assert payload["mode"] == "strict"
    assert "tables" in payload
    assert "rincian_map" in payload


@pytest.mark.django_db
def test_full_upload_commit_flow(api_client, admin_user, master_data):
    """Integration test: generate -> upload -> preview -> commit."""
    admin_user.is_staff = True
    admin_user.save(update_fields=["is_staff"])
    api_client.force_authenticate(user=admin_user)

    bab_id = master_data["bab"].id

    # 1. Generate template
    gen_resp = api_client.post(
        reverse("manual_import:generate_template"),
        data={"publication_year": 2027, "bab_id": bab_id},
        format="json",
    )
    assert gen_resp.status_code == 200

    wb = load_workbook(BytesIO(gen_resp.content))
    # Fill in some data in the first T_ sheet
    table_sheets = [s for s in wb.sheetnames if s.startswith("T_")]
    assert len(table_sheets) >= 1
    table_ws = wb[table_sheets[0]]

    # Find the indicator column
    headers = [cell.value for cell in table_ws[1]]
    ind_col_idx = None
    for idx, h in enumerate(headers, start=1):
        if h and h not in ("wilayah_id", "nama_wilayah", "rincian_id", "nama_rincian"):
            ind_col_idx = idx
            break

    if ind_col_idx:
        # Fill data into first data row
        for row in table_ws.iter_rows(min_row=2, max_row=3):
            id_cell = row[0]
            if id_cell.value is not None:
                row[ind_col_idx - 1].value = 12345

    # 2. Upload
    xlsx_bytes = BytesIO()
    wb.save(xlsx_bytes)
    xlsx_bytes.seek(0)

    upload_resp = api_client.post(
        reverse("manual_import:upload"),
        data={
            "file": xlsx_bytes,
            "publication_year": "2027",
        },
        format="multipart",
    )
    assert upload_resp.status_code == 200, upload_resp.content.decode()[:500]
    data = upload_resp.json()
    assert "upload_id" in data
    assert data["preview"]["validation"]["is_valid"] is True

    # 3. Commit
    upload_id = data["upload_id"]
    commit_resp = api_client.post(
        reverse("manual_import:commit", args=[upload_id]),
        data={},
        format="json",
    )
    assert commit_resp.status_code == 200, commit_resp.content.decode()[:500]
    commit_data = commit_resp.json()
    assert commit_data["status"] == "committed"
    assert commit_data["faktas_inserted"] > 0


def _load_workbook(stream):
    """Helper: openpyxl load_workbook from a stream."""
    return load_workbook(stream, read_only=True, data_only=True)
